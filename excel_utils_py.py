import os
import uuid
import tempfile
from typing import List, Dict, Any
from openpyxl import load_workbook, Workbook
from copy import copy

def guardar_archivo_temporal(upload_file) -> str:
    suffix = os.path.splitext(upload_file.filename)[1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(upload_file.file.read())
        return tmp.name

def extraer_celdas_con_texto(filepath: str) -> List[Dict[str, Any]]:
    wb = load_workbook(filepath)
    ws = wb.active
    resultado = []

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip():
                resultado.append({
                    "row": cell.row,
                    "col": cell.column,
                    "texto": cell.value.strip()
                })

    return resultado

def escribir_excel_traducido(filepath: str, traducciones: List[Dict[str, Any]]) -> str:
    wb = load_workbook(filepath)
    ws_origen = wb.active
    ws_nueva = wb.copy_worksheet(ws_origen)
    ws_nueva.title = "Traducido"

    for t in traducciones:
        cell = ws_nueva.cell(row=t["row"], column=t["col"])
        cell.value = t["texto"]

    salida = os.path.join(tempfile.gettempdir(), f"excel_traducido_{uuid.uuid4()}.xlsx")
    wb.save(salida)
    return salida
